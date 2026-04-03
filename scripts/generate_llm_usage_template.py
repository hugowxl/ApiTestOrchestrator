"""生成大模型相关 Excel 模板：月度资源统计、预算申报。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _styles():
    thin = Side(style="thin", color="CCCCCC")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_fill": PatternFill("solid", fgColor="4472C4"),
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "body_font": Font(size=11),
    }


def write_usage_template(root: Path) -> Path:
    out = root / "docs" / "大模型资源使用_月度统计模板.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    s = _styles()
    border, header_fill, header_font, body_font = (
        s["border"],
        s["header_fill"],
        s["header_font"],
        s["body_font"],
    )

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "使用说明"
    instructions = [
        ("大模型资源使用 — 月度统计模板", True),
        ("", False),
        ("一、用途", False),
        ("  供各部门按月汇总多名用户的大模型调用与费用，便于成本核算与预算管理。", False),
        ("", False),
        ("二、如何填写", False),
        ("  1. 在「月度填报」工作表中按行填写；一名用户若使用多家供应商或多套账号，可拆成多行。", False),
        ("  2. 「统计年月」统一填当月，格式建议 YYYY-MM（如 2026-04）。", False),
        ("  3. Token、调用次数等以各供应商控制台/账单导出为准；若无细分，可只在「本月费用」填总额并在备注说明。", False),
        ("  4. 费用请填人民币或注明币种；包月/资源包可按当月摊销金额填写。", False),
        ("", False),
        ("三、工作表说明", False),
        ("  · 月度填报：主表，每人每供应商（或每项目）每月一行或多行。", False),
        ("  · 字段说明：各列含义、填写规则与示例。", False),
        ("  · 供应商选项：常见供应商名称参考（可复制到填报表使用）。", False),
    ]
    for r, (text, is_title) in enumerate(instructions, 1):
        c = ws0.cell(row=r, column=1, value=text)
        c.font = Font(bold=is_title, size=14 if is_title else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 92

    ws = wb.create_sheet("月度填报", 1)
    headers = [
        "序号",
        "统计年月",
        "用户姓名",
        "用户账号/工号",
        "部门",
        "供应商",
        "子账号或项目ID",
        "产品/控制台名称",
        "主要使用模型",
        "计费方式",
        "本月输入Token",
        "本月输出Token",
        "本月总Token",
        "API调用次数",
        "本月费用(元)",
        "币种",
        "是否含税",
        "费用依据",
        "备注",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in range(2, 17):
        ws.cell(row=row, column=1, value=row - 1).alignment = Alignment(horizontal="center")
        for c in range(2, len(headers) + 1):
            ws.cell(row=row, column=c).border = border

    widths = [6, 12, 10, 14, 12, 14, 18, 16, 18, 12, 14, 14, 14, 12, 14, 8, 10, 14, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    wfd = wb.create_sheet("字段说明", 2)
    fd_headers = ["字段名", "说明", "是否必填", "填写示例"]
    for col, h in enumerate(fd_headers, 1):
        c = wfd.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    rows_fd = [
        ("序号", "行号，便于核对；可自动生成或手填。", "否", "1"),
        ("统计年月", "本次统计对应的自然月。", "是", "2026-04"),
        ("用户姓名", "实际使用大模型的人员。", "是", "张三"),
        ("用户账号/工号", "内部唯一标识，便于与 HR/IT 对账。", "建议填", "w001234 / E12345"),
        ("部门", "所属组织。", "建议填", "研发中心"),
        ("供应商", "大模型或 API 提供方。", "是", "OpenAI / Azure OpenAI / 阿里云百炼"),
        ("子账号或项目ID", "供应商侧账号、资源组、项目 ID 等。", "否", "proj-xxx"),
        ("产品/控制台名称", "计费单元或产品名（与账单一致）。", "否", "Azure OpenAI Service"),
        ("主要使用模型", "用量占比最高的模型，多个可写「多模型」并备注。", "建议填", "gpt-4o"),
        ("计费方式", "按 Token、按次、包月、资源包、混合等。", "建议填", "按 Token"),
        ("本月输入Token", "输入侧 Token 数；无拆分可留空。", "否", "12000000"),
        ("本月输出Token", "输出侧 Token 数。", "否", "8000000"),
        ("本月总Token", "若平台只给合计，可只填本列。", "否", "20000000"),
        ("API调用次数", "请求次数；按次计费时可重点填。", "否", "15000"),
        ("本月费用(元)", "归属到该用户/该行的当月金额（已摊销则填摊销额）。", "是", "1280.50"),
        ("币种", "默认人民币。", "建议填", "CNY"),
        ("是否含税", "与财务口径一致。", "否", "是 / 否"),
        ("费用依据", "发票、对账单、平台导出文件名或链接说明。", "建议填", "202604账单.pdf"),
        ("备注", "分摊规则、测试环境、共享密钥说明等。", "否", "与某项目共用密钥，按调用量比例分摊"),
    ]
    for r, row in enumerate(rows_fd, 2):
        for col, val in enumerate(row, 1):
            cell = wfd.cell(row=r, column=col, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wfd.column_dimensions["A"].width = 18
    wfd.column_dimensions["B"].width = 52
    wfd.column_dimensions["C"].width = 12
    wfd.column_dimensions["D"].width = 28
    wfd.freeze_panes = "A2"

    wsv = wb.create_sheet("供应商选项", 3)
    wsv["A1"] = "常见供应商（填报时「供应商」列可从中选取或自行补充）"
    wsv["A1"].font = Font(bold=True, size=11)
    vendors = [
        "OpenAI",
        "Azure OpenAI",
        "Google (Gemini / Vertex AI)",
        "AWS Bedrock",
        "阿里云百炼 / 通义",
        "腾讯云混元",
        "百度千帆 / 文心",
        "字节火山引擎",
        "智谱 AI (GLM)",
        "MiniMax",
        "月之暗面 (Kimi)",
        "DeepSeek",
        "讯飞星火",
        "华为云盘古",
        "其他（请在备注中写明全称）",
    ]
    for i, v in enumerate(vendors, 3):
        wsv.cell(row=i, column=1, value=v)
    wsv.column_dimensions["A"].width = 40

    wb.save(out)
    return out


def write_budget_template(root: Path) -> Path:
    out = root / "docs" / "大模型预算申报模板.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    s = _styles()
    border, header_fill, header_font, body_font = (
        s["border"],
        s["header_fill"],
        s["header_font"],
        s["body_font"],
    )

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "使用说明"
    instructions = [
        ("大模型服务 — 预算申报模板（精简字段）", True),
        ("", False),
        ("一、用途", False),
        ("  用于在年度、半年度或季度前，汇总各部门拟使用大模型 API 的预算需求，便于统一评审与下达。", False),
        ("", False),
        ("二、如何填写", False),
        ("  1. 在「预算申报」表中按行填写，每一行表示一条预算项（可按项目、按部门汇总或按供应商拆分）。", False),
        ("  2. 「预算周期」请写清口径，例如：2026 年度、2026-H1、2026-Q2；若金额是「月均」请在备注中说明。", False),
        ("  3. 「预算金额」填该周期内预计发生的费用总额（与财务预算口径一致，含税与否在备注说明）。", False),
        ("  4. 拟用供应商或模型尚未敲定时，可写「待定」并在用途说明中描述场景。", False),
        ("", False),
        ("三、工作表", False),
        ("  · 预算申报：主表。", False),
        ("  · 字段说明：各列含义与示例。", False),
    ]
    for r, (text, is_title) in enumerate(instructions, 1):
        c = ws0.cell(row=r, column=1, value=text)
        c.font = Font(bold=is_title, size=14 if is_title else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 88

    ws = wb.create_sheet("预算申报", 1)
    headers = [
        "序号",
        "预算周期",
        "部门",
        "申请人",
        "用途说明",
        "拟用供应商",
        "预算金额(元)",
        "币种",
        "备注",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in range(2, 14):
        ws.cell(row=row, column=1, value=row - 1).alignment = Alignment(horizontal="center")
        for c in range(2, len(headers) + 1):
            ws.cell(row=row, column=c).border = border

    widths = [6, 14, 14, 10, 36, 18, 14, 8, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    wfd = wb.create_sheet("字段说明", 2)
    fd_headers = ["字段名", "说明", "是否必填", "填写示例"]
    for col, h in enumerate(fd_headers, 1):
        c = wfd.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    rows_fd = [
        ("序号", "行号，便于核对。", "否", "1"),
        ("预算周期", "预算覆盖的时间段及口径。", "是", "2026 年度 / 2026-Q2"),
        ("部门", "费用归属部门。", "是", "研发中心"),
        ("申请人", "预算对接人或负责人。", "建议填", "李四"),
        ("用途说明", "业务场景、预期功能、是否对客等，便于评审。", "是", "客服助手知识问答，预计日调用约 5 万次"),
        ("拟用供应商", "计划采购或绑定的 API 提供方；未定可写「待定」。", "建议填", "阿里云百炼"),
        ("预算金额(元)", "该周期内预计总费用（与上面「预算周期」一致）。", "是", "50000"),
        ("币种", "默认人民币。", "建议填", "CNY"),
        ("备注", "含税/不含税、是否月均、是否含测试环境等补充说明。", "否", "含税；金额为全年合计"),
    ]
    for r, row in enumerate(rows_fd, 2):
        for col, val in enumerate(row, 1):
            cell = wfd.cell(row=r, column=col, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wfd.column_dimensions["A"].width = 16
    wfd.column_dimensions["B"].width = 50
    wfd.column_dimensions["C"].width = 12
    wfd.column_dimensions["D"].width = 30
    wfd.freeze_panes = "A2"

    wb.save(out)
    return out


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    p1 = write_usage_template(root)
    p2 = write_budget_template(root)
    print(p1)
    print(p2)


if __name__ == "__main__":
    main()
